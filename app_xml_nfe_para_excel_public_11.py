# app_xml_nfe_para_excel_public.py
# Como usar localmente:
#   1) pip install streamlit pandas openpyxl
#   2) streamlit run app_xml_nfe_para_excel_public.py

import io
import os
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import pandas as pd
import streamlit as st

COLUNAS_NFE = [
    'pasta', 'arquivo_xml', 'tipo_documento', 'chave_acesso', 'modelo', 'serie', 'numero_nf',
    'data_emissao', 'data_saida_entrada', 'tipo_operacao', 'natureza_operacao', 'finalidade',
    'ambiente', 'status_protocolo', 'protocolo_autorizacao', 'emitente_cnpj_cpf',
    'emitente_razao_social', 'emitente_nome_fantasia', 'emitente_ie', 'emitente_crt',
    'origem_uf', 'origem_municipio', 'origem_pais', 'destinatario_cnpj_cpf',
    'destinatario_razao_social', 'destinatario_ie', 'destino_uf', 'destino_municipio',
    'destino_pais', 'n_item', 'codigo_produto', 'descricao_produto', 'ean', 'ncm', 'cest',
    'cfop', 'unidade_comercial', 'quantidade_comercial', 'valor_unitario_comercial',
    'valor_produto', 'unidade_tributavel', 'quantidade_tributavel', 'valor_unitario_tributavel',
    'frete_item', 'seguro_item', 'desconto_item', 'outras_despesas_item',
    'origem_mercadoria_icms', 'icms_cst_csosn', 'bc_icms', 'aliquota_icms', 'valor_icms',
    'bc_icms_st', 'valor_icms_st', 'ipi_cst', 'bc_ipi', 'aliquota_ipi', 'valor_ipi',
    'pis_cst', 'bc_pis', 'aliquota_pis', 'valor_pis', 'cofins_cst', 'bc_cofins',
    'aliquota_cofins', 'valor_cofins', 'valor_total_produtos_nf', 'valor_total_nf',
    'valor_frete_nf', 'valor_seguro_nf', 'valor_desconto_nf', 'valor_outras_despesas_nf',
    'valor_icms_total_nf', 'valor_ipi_total_nf', 'valor_pis_total_nf', 'valor_cofins_total_nf',
    'transportador_cnpj_cpf', 'transportador_nome', 'modalidade_frete', 'placa_veiculo',
    'uf_veiculo', 'peso_bruto', 'peso_liquido', 'informacoes_complementares'
]

COLUNAS_EVENTOS = [
    'pasta', 'arquivo_xml', 'tipo_evento', 'descricao_evento',
    'chave_nfe', 'cnpj_autor', 'ambiente',
    'data_evento', 'data_registro',
    'numero_protocolo', 'status_codigo', 'status_descricao',
    'justificativa', 'texto_correcao',
]

TIPOS_EVENTO = {
    "110111": "Cancelamento",
    "110110": "Carta de Correção",
    "110112": "Cancelamento por Substituição",
    "110114": "EPEC",
    "210200": "Ciência da Operação",
    "210210": "Confirmação da Operação",
    "210220": "Desconhecimento da Operação",
    "210240": "Operação não Realizada",
    "610130": "Comprovante de Entrega CT-e",
}


def limpar_namespace(root):
    for elem in root.iter():
        if "}" in elem.tag:
            elem.tag = elem.tag.split("}", 1)[1]
    return root


def txt(parent, path, default=""):
    if parent is None:
        return default
    found = parent.find(path)
    if found is None or found.text is None:
        return default
    return found.text.strip()


def attr(parent, name, default=""):
    if parent is None:
        return default
    return parent.attrib.get(name, default)


def num(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return None


def cnpj_ou_cpf(parent):
    return txt(parent, "CNPJ") or txt(parent, "CPF") or txt(parent, "idEstrangeiro")


def primeiro_filho(parent):
    if parent is None:
        return None
    filhos = list(parent)
    return filhos[0] if filhos else None


def detectar_tipo(modelo):
    if modelo == "55": return "NF-e"
    if modelo == "65": return "NFC-e"
    if modelo in {"57", "67"}: return "CT-e/CT-e OS"
    return "XML fiscal"


def parse_evento(root, nome_arquivo, pasta):
    linhas = []
    inf = root.find(".//infEvento")
    if inf is None:
        return linhas

    tipo_cod = txt(inf, "tpEvento")
    tipo_desc = TIPOS_EVENTO.get(tipo_cod, f"Evento {tipo_cod}")
    det = inf.find("detEvento")
    desc_evento = txt(det, "descEvento") if det is not None else tipo_desc
    justificativa = txt(det, "xJust") if det is not None else ""
    texto_correcao = txt(det, "xCorrecao") if det is not None else ""
    ambiente = {"1": "Produção", "2": "Homologação"}.get(txt(inf, "tpAmb"), txt(inf, "tpAmb"))

    ret = root.find(".//retEvento/infEvento")

    linhas.append({
        "pasta": pasta,
        "arquivo_xml": nome_arquivo,
        "tipo_evento": tipo_desc,
        "descricao_evento": desc_evento,
        "chave_nfe": txt(inf, "chNFe"),
        "cnpj_autor": cnpj_ou_cpf(inf),
        "ambiente": ambiente,
        "data_evento": txt(inf, "dhEvento"),
        "data_registro": txt(ret, "dhRegEvento") if ret is not None else "",
        "numero_protocolo": txt(ret, "nProt") if ret is not None else "",
        "status_codigo": txt(ret, "cStat") if ret is not None else "",
        "status_descricao": txt(ret, "xMotivo") if ret is not None else "",
        "justificativa": justificativa,
        "texto_correcao": texto_correcao,
    })
    return linhas


def parse_xml_nfe(xml_bytes, nome_arquivo, pasta):
    """Retorna (linhas_nfe, linhas_eventos, linhas_erros)"""
    try:
        root = ET.fromstring(xml_bytes)
        root = limpar_namespace(root)
    except Exception as e:
        return [], [], [{"pasta": pasta, "arquivo_xml": nome_arquivo, "erro": f"XML inválido: {e}"}]

    # É um evento?
    if root.find(".//infEvento") is not None:
        return [], parse_evento(root, nome_arquivo, pasta), []

    # É uma NF-e?
    inf_nfe = root.find(".//infNFe")
    if inf_nfe is None:
        return [], [], [{"pasta": pasta, "arquivo_xml": nome_arquivo, "erro": "Formato não reconhecido."}]

    ide = inf_nfe.find("ide")
    emit = inf_nfe.find("emit")
    dest = inf_nfe.find("dest")
    end_emit = emit.find("enderEmit") if emit is not None else None
    end_dest = dest.find("enderDest") if dest is not None else None
    total = inf_nfe.find("total/ICMSTot")
    transp = inf_nfe.find("transp")
    transporta = transp.find("transporta") if transp is not None else None
    veic = transp.find("veicTransp") if transp is not None else None
    vol = transp.find("vol") if transp is not None else None
    inf_prot = root.find(".//protNFe/infProt")

    chave = attr(inf_nfe, "Id").replace("NFe", "") or txt(inf_prot, "chNFe")
    modelo = txt(ide, "mod")
    tp_nf_desc = {"0": "Entrada", "1": "Saída"}.get(txt(ide, "tpNF"), txt(ide, "tpNF"))
    ambiente = {"1": "Produção", "2": "Homologação"}.get(txt(ide, "tpAmb"), txt(ide, "tpAmb"))
    finalidade = {
        "1": "NF-e normal", "2": "NF-e complementar",
        "3": "NF-e de ajuste", "4": "Devolução/retorno",
    }.get(txt(ide, "finNFe"), txt(ide, "finNFe"))
    mod_frete_val = txt(transp, "modFrete") if transp is not None else ""
    mod_frete = {
        "0": "Por conta do emitente",
        "1": "Por conta do destinatário/remetente",
        "2": "Por conta de terceiros",
        "3": "Transporte próprio por conta do remetente",
        "4": "Transporte próprio por conta do destinatário",
        "9": "Sem frete",
    }.get(mod_frete_val, mod_frete_val)

    capa = {
        "pasta": pasta,
        "arquivo_xml": nome_arquivo,
        "tipo_documento": detectar_tipo(modelo),
        "chave_acesso": chave,
        "modelo": modelo,
        "serie": txt(ide, "serie"),
        "numero_nf": txt(ide, "nNF"),
        "data_emissao": txt(ide, "dhEmi") or txt(ide, "dEmi"),
        "data_saida_entrada": txt(ide, "dhSaiEnt") or txt(ide, "dSaiEnt"),
        "tipo_operacao": tp_nf_desc,
        "natureza_operacao": txt(ide, "natOp"),
        "finalidade": finalidade,
        "ambiente": ambiente,
        "status_protocolo": txt(inf_prot, "cStat"),
        "protocolo_autorizacao": txt(inf_prot, "nProt"),
        "emitente_cnpj_cpf": cnpj_ou_cpf(emit),
        "emitente_razao_social": txt(emit, "xNome"),
        "emitente_nome_fantasia": txt(emit, "xFant"),
        "emitente_ie": txt(emit, "IE"),
        "emitente_crt": txt(emit, "CRT"),
        "origem_uf": txt(end_emit, "UF"),
        "origem_municipio": txt(end_emit, "xMun"),
        "origem_pais": txt(end_emit, "xPais"),
        "destinatario_cnpj_cpf": cnpj_ou_cpf(dest),
        "destinatario_razao_social": txt(dest, "xNome"),
        "destinatario_ie": txt(dest, "IE"),
        "destino_uf": txt(end_dest, "UF"),
        "destino_municipio": txt(end_dest, "xMun"),
        "destino_pais": txt(end_dest, "xPais"),
        "valor_total_produtos_nf": num(txt(total, "vProd")),
        "valor_total_nf": num(txt(total, "vNF")),
        "valor_frete_nf": num(txt(total, "vFrete")),
        "valor_seguro_nf": num(txt(total, "vSeg")),
        "valor_desconto_nf": num(txt(total, "vDesc")),
        "valor_outras_despesas_nf": num(txt(total, "vOutro")),
        "valor_icms_total_nf": num(txt(total, "vICMS")),
        "valor_ipi_total_nf": num(txt(total, "vIPI")),
        "valor_pis_total_nf": num(txt(total, "vPIS")),
        "valor_cofins_total_nf": num(txt(total, "vCOFINS")),
        "transportador_cnpj_cpf": cnpj_ou_cpf(transporta),
        "transportador_nome": txt(transporta, "xNome"),
        "modalidade_frete": mod_frete,
        "placa_veiculo": txt(veic, "placa"),
        "uf_veiculo": txt(veic, "UF"),
        "peso_bruto": num(txt(vol, "pesoB")),
        "peso_liquido": num(txt(vol, "pesoL")),
        "informacoes_complementares": txt(inf_nfe.find("infAdic"), "infCpl"),
    }

    detalhes = inf_nfe.findall("det")
    if not detalhes:
        linha = {col: "" for col in COLUNAS_NFE}
        linha.update(capa)
        return [linha], [], []

    linhas = []
    for det in detalhes:
        prod = det.find("prod")
        imposto = det.find("imposto")
        icms_grupo = primeiro_filho(imposto.find("ICMS") if imposto is not None else None)
        ipi = imposto.find("IPI") if imposto is not None else None
        ipi_grupo = primeiro_filho(ipi) if ipi is not None else None
        pis_grupo = primeiro_filho(imposto.find("PIS") if imposto is not None else None)
        cofins_grupo = primeiro_filho(imposto.find("COFINS") if imposto is not None else None)

        linha = {col: "" for col in COLUNAS_NFE}
        linha.update(capa)
        linha.update({
            "n_item": attr(det, "nItem"),
            "codigo_produto": txt(prod, "cProd"),
            "descricao_produto": txt(prod, "xProd"),
            "ean": txt(prod, "cEAN"),
            "ncm": txt(prod, "NCM"),
            "cest": txt(prod, "CEST"),
            "cfop": txt(prod, "CFOP"),
            "unidade_comercial": txt(prod, "uCom"),
            "quantidade_comercial": num(txt(prod, "qCom")),
            "valor_unitario_comercial": num(txt(prod, "vUnCom")),
            "valor_produto": num(txt(prod, "vProd")),
            "unidade_tributavel": txt(prod, "uTrib"),
            "quantidade_tributavel": num(txt(prod, "qTrib")),
            "valor_unitario_tributavel": num(txt(prod, "vUnTrib")),
            "frete_item": num(txt(prod, "vFrete")),
            "seguro_item": num(txt(prod, "vSeg")),
            "desconto_item": num(txt(prod, "vDesc")),
            "outras_despesas_item": num(txt(prod, "vOutro")),
            "origem_mercadoria_icms": txt(icms_grupo, "orig"),
            "icms_cst_csosn": txt(icms_grupo, "CST") or txt(icms_grupo, "CSOSN"),
            "bc_icms": num(txt(icms_grupo, "vBC")),
            "aliquota_icms": num(txt(icms_grupo, "pICMS")),
            "valor_icms": num(txt(icms_grupo, "vICMS")),
            "bc_icms_st": num(txt(icms_grupo, "vBCST")),
            "valor_icms_st": num(txt(icms_grupo, "vICMSST")),
            "ipi_cst": txt(ipi_grupo, "CST"),
            "bc_ipi": num(txt(ipi_grupo, "vBC")),
            "aliquota_ipi": num(txt(ipi_grupo, "pIPI")),
            "valor_ipi": num(txt(ipi_grupo, "vIPI")),
            "pis_cst": txt(pis_grupo, "CST"),
            "bc_pis": num(txt(pis_grupo, "vBC")),
            "aliquota_pis": num(txt(pis_grupo, "pPIS")),
            "valor_pis": num(txt(pis_grupo, "vPIS")),
            "cofins_cst": txt(cofins_grupo, "CST"),
            "bc_cofins": num(txt(cofins_grupo, "vBC")),
            "aliquota_cofins": num(txt(cofins_grupo, "pCOFINS")),
            "valor_cofins": num(txt(cofins_grupo, "vCOFINS")),
        })
        linhas.append(linha)
    return linhas, [], []


def ler_uploads(uploaded_files):
    arquivos = []
    for up in uploaded_files:
        data = up.read()
        nome = up.name
        if nome.lower().endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                for item in zf.namelist():
                    if item.lower().endswith(".xml"):
                        conteudo = zf.read(item)
                        pasta = os.path.dirname(item) or "raiz"
                        arquivos.append((pasta, os.path.basename(item), conteudo))
        elif nome.lower().endswith(".xml"):
            arquivos.append(("avulso",nome, data))
    return arquivos
      


def estilo_aba(ws, cor="1F4E78"):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        header = str(col_cells[0].value or "")
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[:200])
        width = min(max(max_len + 2, 12), 42)
        if any(x in header for x in ["descricao", "razao", "informacoes", "natureza", "justificativa", "correcao", "status"]):
            width = 38
        ws.column_dimensions[col_cells[0].column_letter].width = width
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True, color="FFFFFF")
        cell.fill = cell.fill.copy(fill_type="solid", fgColor=cor)


def colorir_duplicatas(ws):
    """Pinta de vermelho claro as linhas onde coluna duplicada == Sim"""
    from openpyxl.styles import PatternFill
    fill_dup = PatternFill(fill_type="solid", fgColor="FFCCCC")
    col_dup = None
    for cell in ws[1]:
        if cell.value == "duplicada":
            col_dup = cell.column
            break
    if col_dup is None:
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[col_dup - 1].value == "Sim":
            for cell in row:
                cell.fill = fill_dup


def gerar_excel(df_nfe, df_eventos, df_erros):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_nfe.to_excel(writer, index=False, sheet_name="Notas_Fiscais")
        if not df_eventos.empty:
            df_eventos.to_excel(writer, index=False, sheet_name="Eventos_Cancelamentos")
        if not df_erros.empty:
            df_erros.to_excel(writer, index=False, sheet_name="Erros")
        wb = writer.book
        estilo_aba(wb["Notas_Fiscais"], "1F4E78")
        colorir_duplicatas(wb["Notas_Fiscais"])
        if not df_eventos.empty:
            estilo_aba(wb["Eventos_Cancelamentos"], "7B2D00")
            colorir_duplicatas(wb["Eventos_Cancelamentos"])
        if not df_erros.empty:
            estilo_aba(wb["Erros"], "5C0000")
    output.seek(0)
    return output


# ── Interface ──────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Conversor XML NF-e para Excel - Countout Co.", layout="wide")
st.title("📄 Conversor de XML de notas fiscais para Excel - Countout Co.")
st.caption("Envie aqui os arquivos. O relatório sai com uma linha por item da nota."
            "Caso encontre notas duplicadas, apenas filtre a opção SIM e exclua todas as linhas."
           "Dessa maneira ficará apenas um arquivo correto. Bem bunitin e padrão Countout hehe 😁" )

if "upload_key" not in st.session_state:
    st.session_state.upload_key = 0

uploaded = st.file_uploader(
    "Selecione os arquivos (ZIP ou XMLs avulsos)",
    type=["zip", "xml"],
    accept_multiple_files=True,
    key=f"uploader_{st.session_state.upload_key}",
)

if uploaded:
    if st.button("🔄 Limpar e recomeçar"):
        st.session_state.upload_key += 1
        st.rerun()

    with st.spinner("Lendo arquivos..."):
        arquivos = ler_uploads(uploaded)

    if not arquivos:
        st.error("Nenhum arquivo .xml encontrado. Verifique o conteúdo e tente novamente.")
    else:
        progresso = st.progress(0, text="Iniciando...")
        linhas_nfe, linhas_eventos, linhas_erros = [], [], []
        total = len(arquivos)

        for i, (pasta, nome, data) in enumerate(arquivos):
            l, e, err = parse_xml_nfe(data, nome, pasta)
            linhas_nfe.extend(l)
            linhas_eventos.extend(e)
            linhas_erros.extend(err)
            progresso.progress((i + 1) / total, text=f"Processando {i+1}/{total}: {nome[:60]}")

        progresso.empty()

        df_nfe = pd.DataFrame(linhas_nfe, columns=COLUNAS_NFE)
        df_eventos = pd.DataFrame(linhas_eventos, columns=COLUNAS_EVENTOS) if linhas_eventos else pd.DataFrame(columns=COLUNAS_EVENTOS)
        df_erros = pd.DataFrame(linhas_erros)

        # Identificar duplicatas:
        # A chave_acesso é o identificador único fiscal da NF-e (44 dígitos).
        # Combinada com n_item e codigo_produto garante que só é duplicata
        # quando o exato mesmo item da exata mesma nota aparece mais de uma vez.
        # arquivo_xml e numero_nf NÃO são usados: dois XMLs distintos podem ter
        # o mesmo nome de arquivo ou mesmo número de NF de emitentes diferentes.
        if not df_nfe.empty:
            df_nfe.insert(0, "duplicada", df_nfe.duplicated(
                subset=["chave_acesso", "n_item", "codigo_produto"], keep="first"
            ).map({True: "Sim", False: "Não"}))
        duplicadas_nfe = int((df_nfe["duplicada"] == "Sim").sum()) if not df_nfe.empty else 0

        # Eventos: mesma chave + mesmo tipo + mesmo protocolo
        if not df_eventos.empty:
            df_eventos.insert(0, "duplicada", df_eventos.duplicated(subset=["chave_nfe", "tipo_evento", "numero_protocolo"], keep="first").map({True: "Sim", False: "Não"}))
        duplicadas_ev = int((df_eventos["duplicada"] == "Sim").sum()) if not df_eventos.empty else 0

        msg = (
            f"✅ Total: {total} XML(s) | "
            f"Notas fiscais: {len(df_nfe)} linhas | "
            f"Eventos/cancelamentos: {len(df_eventos)} | "
            f"Erros: {len(df_erros)}"
        )
        if duplicadas_nfe > 0 or duplicadas_ev > 0:
            msg += f" | ⚠️ Duplicatas encontradas: {duplicadas_nfe} notas + {duplicadas_ev} eventos (coluna 'duplicada' = Sim)"
        st.success(msg)

        tab1, tab2 = st.tabs(["📋 Notas Fiscais", "🚫 Eventos e Cancelamentos"])

        with tab1:
            st.dataframe(df_nfe, use_container_width=True, height=420)

        with tab2:
            if df_eventos.empty:
                st.info("Nenhum evento ou cancelamento encontrado.")
            else:
                # Resumo por tipo
                resumo = df_eventos.groupby("tipo_evento").size().reset_index(name="quantidade")
                st.dataframe(resumo, use_container_width=True, hide_index=True)
                st.dataframe(df_eventos, use_container_width=True, height=380)

        excel = gerar_excel(df_nfe, df_eventos, df_erros)
        st.download_button(
            "⬇️ Baixar relatório Excel (todas as abas).",
            data=excel,
            file_name=f"relatorio_xml_nfe_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
